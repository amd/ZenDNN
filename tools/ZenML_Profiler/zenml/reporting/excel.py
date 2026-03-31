# *******************************************************************************
# Copyright (c) 2026 Advanced Micro Devices, Inc. All rights reserved.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.
#
# *******************************************************************************
import logging
import os
import re
import pytz
import time as time1
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.chart import PieChart, LineChart, Reference, Series
from zenml.reporting.utils import non_zero_round

_EXCEL_OUTPUT_DIR = "./Excel_Analysis"
_REPORT_TIMEZONE = "Asia/Kolkata"


def cell_insertion(ws, ls, start_row=1, col=4):
    """
    Inserts values into the worksheet using cell number

    Parameters:
    ws (Workbook) : contains the worksheet on which values needs to
                    be inserted
    ls (list) : list of values to be inserted

    Optional parameters:
    start_row (int) : denotes the starting row number, where
                      values needs to be inserted
    col (int) :  denotes the starting column number, where
                 values needs to be inserted
    """
    for index, value in enumerate(ls):
        current_row = start_row + index
        ws.cell(row=current_row, column=col, value=value)


def excel_generator(args, backends, t_time, group_summary, op_summary, flops, comp=None, emb_dfs=None):
    """
    Creates excel view for all the tables created. Charts for each type
    of summary for all logs are also created and attached to the excel.

    Parameters:
    args (ArgParser) : Argument parser object
    backends (list of strings) : contains all backends of logs
    t_time (string) : total time for each iteration
    group_summary (list of data frames) : contains all group summary data frames
    op_summary (list of data frames) : contains all detailed op summary data frames
    flops (list of data frames) : contains all flop summary data frames

    Optional parameters:
    comp (list) : contains comparison table information
    emb_dfs (list) : contains embedding summary data frames
    """
    if comp is None:
        comp = []
    if emb_dfs is None:
        emb_dfs = []
    directory = _EXCEL_OUTPUT_DIR
    # if folder doesnt exist, create one
    if not os.path.exists(directory):
        os.makedirs(directory)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Grouped summary"
    ws2 = wb.create_sheet(title="Detailed summary")
    ws3 = wb.create_sheet(title="Efficiency")
    length = []
    zipped = (
        zip([group_summary, op_summary, flops], [ws1, ws2, ws3])
        if len(flops) > 0
        else zip([group_summary, op_summary], [ws1, ws2])
    )
    temp_count = 0
    for dfs, ws in zipped:
        length.append([])
        for en, df in enumerate(dfs):
            length[-1].append(len(df))
            tmp_col = ["S.No"]
            tmp_col.extend(df.columns.tolist())
            ws.append(tmp_col)
            row_sno = 1
            for _, row in df.iterrows():
                tmp_row = [row_sno]
                tmp_row.extend(row.tolist())
                ws.append(tmp_row)
                row_sno += 1
            if temp_count != 2:
                for _ in range(17):
                    ws.append([])
            else:
                ws.append([])
                ws.append([])
        temp_count += 1

    if len(comp) != 0:
        ws_c = wb.create_sheet(title="Op Time Comparison")
        comp_table, comp_header, comp_rows, mm_result = comp[0]
        header_row = ["S.No"]
        for h in comp_header:
            header_row.append(h.replace("\n", " "))
        ws_c.append(header_row)
        for row_sno, row in enumerate(comp_rows, start=1):
            temp_row = [row_sno]
            for cell in row:
                if isinstance(cell, str):
                    temp_row.append(cell.replace("\n", ""))
                else:
                    temp_row.append(cell)
            ws_c.append(temp_row)

        if mm_result is not None:
            ws_mm = wb.create_sheet(title="Matmul Comparison")
            mm_table, mm_header, mm_rows = mm_result
            mm_header_row = ["S.No"]
            for h in mm_header:
                mm_header_row.append(h.replace("\n", " ") if isinstance(h, str) else h)
            ws_mm.append(mm_header_row)
            for row_sno, row in enumerate(mm_rows, start=1):
                temp_row = [row_sno]
                for cell in row:
                    if isinstance(cell, str):
                        temp_row.append(cell.replace("\n", ""))
                    else:
                        temp_row.append(cell)
                ws_mm.append(temp_row)
    ws4 = wb.create_sheet(title="Chart Extras")
    grp = []
    zipped = (
        zip(range(3), [group_summary, op_summary, flops])
        if isinstance(flops, list)
        else zip(range(2), [group_summary, op_summary])
    )
    for i, dfs in zipped:
        c = 0
        for j, df in enumerate(dfs):
            if isinstance(getattr(args, f"log_name_{j+1}"), bool):
                ch_name = f"Log {j+1}"
            else:
                ch_name = getattr(args, f"log_name_{j+1}")
            if i == 0:
                group_pie = {"op_name": [], "time": []}
                temp_per = 0
                flag = True
                df_count = 0
                for time, op in zip(
                    (
                        df["% E2E Impact"]
                        if t_time[j] != ""
                        else df["% Primitive Exec Impact"]
                    ),
                    df["Op Type"],
                ):
                    time = float(time)
                    if temp_per < 90 and time > 1 and df_count < 9:
                        temp_per += time
                        group_pie["op_name"].append(op)
                        group_pie["time"].append(non_zero_round(time, 2))
                        df_count += 1
                    elif time <= 1:
                        flag = False
                    else:
                        flag = False
                        break
                if not flag or non_zero_round(temp_per) != 100:
                    group_pie["op_name"].append("Others")
                    group_pie["time"].append(non_zero_round(100 - temp_per, 2))
                    df_count += 1
                temp_df = pd.DataFrame(
                    {
                        "Op Name": group_pie["op_name"],
                        ch_name + " Time": group_pie["time"],
                    }
                )
                ws4.append(temp_df.columns.tolist())
                for _, row in temp_df.iterrows():
                    ws4.append(row.tolist())
                ws4.append([])
                pie_chart = PieChart()
                pie_chart.title = ch_name + " Op Distribution"
                labels = Reference(
                    ws4, min_col=1, min_row=c + 2, max_row=c + df_count + 1
                )
                data = Reference(
                    ws4, min_col=2, min_row=c + 1, max_row=c + df_count + 1
                )
                pie_chart.add_data(data, titles_from_data=True)
                pie_chart.set_categories(labels)

                ws1.add_chart(
                    pie_chart,
                    "A"
                    + str((17 * j) + sum([ls for ls in length[0][: j + 1]]) + j + 3),
                )
                grp.append(group_pie)
                if c == 0:
                    c = df_count + 2
                else:
                    c += df_count + 2
            if i == 1:
                new_dim = []
                for op_name, dim in zip(df["Op Type"], df["Dimension"]):
                    if op_name in ["matmul", "batch_matmul", "gemm_api"]:
                        m = dim.split(":")[0].split("x")[-2]
                        n = dim.split(":")[1].split("x")[-1]
                        k = dim.split(":")[0].split("x")[-1]
                        if op_name == "batch_matmul":
                            batches = dim.split(":")[0].split("x")[0]
                        else:
                            batches = "1"
                        new_dim.append(
                            "B:" + batches + ",M:" + m + ",N:" + n + ",K:" + k
                        )
                    elif op_name == "convolution":
                        conv_match = re.match(
                            r"mb(\d+)(?:_?)ic(\d+)oc(\d+)_(.+)_(.+)", dim
                        )
                        if conv_match:
                            m = conv_match.group(1)
                            k = conv_match.group(2)
                            n = conv_match.group(3)
                            h = conv_match.group(4)
                            w = conv_match.group(5)
                            new_dim.append(
                                f"M:{m},N:{n},K:{k},height dimension:{h},width dimension:{w}"
                            )
                        else:
                            new_dim.append(dim)
                    else:
                        new_dim.append(dim)
                df["Dimension"] = new_dim
                top_3 = 1
                for category, grouped_df in df.groupby("Op Type"):
                    op_pie = None
                    if category in grp[j]["op_name"]:
                        op_pie = {"op_dim": [], "time": []}
                        temp_per = 0
                        total_sum = sum(grouped_df["Total Time (ms)"])
                        flag = True
                        df_count = 0
                        if total_sum > 0:
                            for time, op in zip(
                                grouped_df["Total Time (ms)"], grouped_df["Dimension"]
                            ):
                                if (
                                    temp_per / total_sum < 0.9
                                    and time / total_sum > 0.01
                                    and df_count < 9
                                ):
                                    temp_per += time
                                    op_pie["op_dim"].append(op)
                                    op_pie["time"].append(non_zero_round(time, 2))
                                    df_count += 1
                                elif time <= 1:
                                    flag = False
                                else:
                                    flag = False
                                    break
                        if total_sum > 0 and (not flag or temp_per / total_sum != 1.00):
                            op_pie["op_dim"].append("other")
                            op_pie["time"].append(
                                non_zero_round(total_sum - temp_per, 2)
                            )
                            df_count += 1
                    if top_3 <= 3 and op_pie is not None:
                        op_pie["op_dim"].insert(0, "Dimension")
                        op_pie["time"].insert(
                            0, ch_name + " " + str(category) + " Time"
                        )
                        row = 1 + (12 * (j))
                        cell_insertion(
                            ws4, op_pie["op_dim"], start_row=row, col=top_3 * 3 + 1
                        )
                        cell_insertion(
                            ws4, op_pie["time"], start_row=row, col=top_3 * 3 + 2
                        )
                        pie_chart = PieChart()
                        pie_chart.title = ch_name + " " + str(category) + " Distribution"
                        labels = Reference(
                            ws4,
                            min_col=top_3 * 3 + 1,
                            min_row=row + 1,
                            max_row=row + df_count,
                        )
                        data = Reference(
                            ws4,
                            min_col=top_3 * 3 + 2,
                            min_row=row,
                            max_row=row + df_count,
                        )
                        pie_chart.add_data(data, titles_from_data=True)
                        pie_chart.set_categories(labels)
                        loc = str(
                            chr((64 + top_3 * 9 - 8))
                            + str(
                                (17 * j)
                                + sum([ls for ls in length[1][: j + 1]])
                                + j
                                + 3
                            )
                        )
                        ws2.add_chart(pie_chart, loc)
                        top_3 += 1
            if i == 2:
                if j == 0:
                    chart1 = LineChart()
                if args.flops:
                    col = len(df.columns) + 1
                else:
                    col = len(df.columns)
                values = Reference(
                    ws3,
                    min_col=col,
                    min_row=(2 + sum([ls for ls in length[2][:j]]) + (3 * j)),
                    max_row=(1 + sum([ls for ls in length[2][:j]]) + (3 * j)) + len(df),
                )
                series = Series(values, title=ch_name)
                chart1.append(series)
                s1 = chart1.series[-1]
                s1.marker.symbol = "circle"
                if args.flops:
                    chart1.y_axis.title = "Theoretical Efficiency"
                else:
                    chart1.y_axis.title = "GFlops"

                chart1.x_axis.title = "S.No"
                chart1.title = "Efficiency Comparison"
                if j == len(backends) - 1:
                    ws3.add_chart(chart1, "Q1")

    if len(comp) != 0:
        comp_table, comp_header, comp_rows, _ = comp[0]
        n_backends = len(backends)
        count = len(comp_rows)

        time_col_indices = [3 * n_backends + i for i in range(n_backends)]
        ratio_col_indices = [6 * n_backends + i for i in range(n_backends - 1)]

        complete_rows = [
            row for row in comp_rows
            if all(row[ci] != "-" for ci in time_col_indices)
        ]

        if complete_rows:
            cd_start = ws4.max_row + 2
            chart_hdr = ["S.No"]
            for ind in range(n_backends):
                chart_hdr.append(comp_header[3 * n_backends + ind].replace("\n", " "))
            for ind in range(n_backends - 1):
                chart_hdr.append(comp_header[6 * n_backends + ind].replace("\n", " "))
            for ci, h in enumerate(chart_hdr):
                ws4.cell(row=cd_start, column=ci + 1, value=h)

            for ri, row in enumerate(complete_rows, start=1):
                data_row = cd_start + ri
                ws4.cell(row=data_row, column=1, value=ri)
                col = 2
                for ind in range(n_backends):
                    ws4.cell(row=data_row, column=col, value=row[time_col_indices[ind]])
                    col += 1
                for ind in range(n_backends - 1):
                    val = row[ratio_col_indices[ind]]
                    ws4.cell(row=data_row, column=col, value=val if val != "-" else None)
                    col += 1

            n_complete = len(complete_rows)
            cd_first = cd_start + 1
            cd_last = cd_start + n_complete
            cats = Reference(ws4, min_col=1, min_row=cd_first, max_row=cd_last)

            chart2 = LineChart()
            chart2.title = "Op Time Comparison"
            for ind in range(n_backends):
                values1 = Reference(ws4, min_col=2 + ind, min_row=cd_first, max_row=cd_last)
                series1 = Series(values1, title=comp_header[3 * n_backends + ind].replace("\n", " "))
                chart2.append(series1)
                chart2.series[-1].marker.symbol = "circle"
            chart2.set_categories(cats)
            chart2.y_axis.title = "Time (ms)"
            chart2.y_axis.delete = False
            chart2.y_axis.tickLblPos = "nextTo"
            chart2.x_axis.title = "S.No"
            chart2.x_axis.delete = False
            chart2.x_axis.tickLblPos = "nextTo"
            chart2.width = 22
            chart2.height = 15
            chart_anchor_row = count + 3
            ws_c.add_chart(chart2, "A" + str(chart_anchor_row))

            n_ratios = n_backends - 1
            if n_ratios > 0:
                chart3 = LineChart()
                chart3.title = "Time Ratio Comparison"
                for ind in range(n_ratios):
                    ratio_col = 2 + n_backends + ind
                    values_r = Reference(ws4, min_col=ratio_col, min_row=cd_first, max_row=cd_last)
                    ratio_title = comp_header[6 * n_backends + ind].replace("\n", " ")
                    series_r = Series(values_r, title=ratio_title)
                    chart3.append(series_r)
                    chart3.series[-1].marker.symbol = "circle"
                chart3.set_categories(cats)
                chart3.y_axis.title = "Ratio"
                chart3.y_axis.delete = False
                chart3.y_axis.tickLblPos = "nextTo"
                chart3.x_axis.title = "S.No"
                chart3.x_axis.delete = False
                chart3.x_axis.tickLblPos = "nextTo"
                chart3.width = 22
                chart3.height = 15
                ws_c.add_chart(chart3, "P" + str(chart_anchor_row))
    if emb_dfs:
        ws_emb = wb.create_sheet(title="Embedding Summary")
        for en, df in enumerate(emb_dfs):
            ws_emb.append(["S.No"] + df.columns.tolist())
            for row_sno, (_, row) in enumerate(df.iterrows(), start=1):
                ws_emb.append([row_sno] + row.tolist())
            ws_emb.append([])

    timestamp = time1.time()
    date_time = datetime.fromtimestamp(timestamp)
    ist = pytz.timezone(_REPORT_TIMEZONE)
    ist_time = date_time.astimezone(ist)
    formatted_date_time = ist_time.strftime("%d-%m-%Y_%H_%M_%S")
    file_name = (
        str(directory)
        + "/"
        + "ZenProfiler_Report_"
        + str(formatted_date_time)
        + ".xlsx"
    )
    wb.save(file_name)
    logging.info("Excel report saved: %s", file_name)
